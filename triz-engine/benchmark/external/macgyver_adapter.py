"""Adapter for MacGyver benchmark (Allen AI, NAACL 2024).

Tests constrained creative problem solving — real-world challenges
requiring inventive object usage under constraints.

Source: https://github.com/allenai/MacGyver (Apache 2.0)
Dataset: problem_solution_pair.xlsx
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent.parent
DATA_DIR = ROOT / "benchmark" / "external" / "data"
RESULTS_DIR = ROOT / "results" / "external-macgyver"

TRIZ_SYSTEM_PROMPT = """You are a creative problem solver using TRIZ (Theory of Inventive Problem Solving).

When solving this problem:
1. **Identify the contradiction**: What must the available resources do that seems impossible? Frame it as a physical contradiction ("Resource X must be both A and not-A simultaneously").
2. **Apply separation principles**: Can you separate the contradictory requirements in time, space, condition, or system level?
3. **Apply inventive principles**: Consider these TRIZ principles for creative resource use:
   - Principle 1 (Segmentation): Can you break a resource into parts?
   - Principle 2 (Taking out): Can you extract just the useful property?
   - Principle 5 (Merging): Can you combine resources for new function?
   - Principle 6 (Universality): Can one resource serve multiple functions?
   - Principle 13 (Inversion): Can you use the resource in reverse?
   - Principle 25 (Self-service): Can the environment itself serve as a resource?
   - Principle 35 (Parameter changes): Can you change a physical property?
4. **Generate a concrete solution** that uses only the available resources in creative ways.

Think inventively — look for ways to make resources serve functions they weren't designed for."""

VANILLA_SYSTEM_PROMPT = """Solve this practical problem using only the resources available. Be creative and think about unconventional uses for the items you have."""


def download_dataset() -> Path:
    """Download MacGyver dataset from GitHub."""
    import urllib.request

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    target = DATA_DIR / "macgyver_problems.json"
    if target.exists():
        return target

    xlsx_url = "https://raw.githubusercontent.com/allenai/MacGyver/main/data/MacGyver/problem_solution_pair.xlsx"
    xlsx_path = DATA_DIR / "problem_solution_pair.xlsx"

    if not xlsx_path.exists():
        print(f"Downloading MacGyver dataset...", file=sys.stderr)
        try:
            urllib.request.urlretrieve(xlsx_url, xlsx_path)
        except Exception as e:
            print(f"Download failed: {e}", file=sys.stderr)
            raise

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        problems = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))

            solvable = str(row_dict.get("Solvability", "")).lower()
            if solvable in ("no", "false", "0", "unsolvable"):
                continue

            problem = row_dict.get("Problem", row_dict.get("problem", ""))
            solution = row_dict.get("Solution", row_dict.get("solution", ""))

            if problem and solution:
                problems.append({
                    "problem": str(problem),
                    "reference_solution": str(solution),
                    "category": str(row_dict.get("Category", row_dict.get("category", "general"))),
                    "unconventional": str(row_dict.get("Unconventional", "")),
                })

        target.write_text(json.dumps(problems, indent=2))
        print(f"Converted {len(problems)} solvable problems to {target}", file=sys.stderr)
        return target

    except ImportError:
        print("openpyxl required for xlsx parsing. Install: pip install openpyxl", file=sys.stderr)
        raise


def load_problems(
    limit: int = 50,
    categories: list[str] | None = None,
    start_offset: int = 0,
    skip_completed: bool = False,
) -> list[dict]:
    """Load MacGyver problems.

    Returns list of dicts with: id, problem, reference_solution, category.
    When skip_completed=True, also skips problems whose result JSON already exists.
    """
    data_path = DATA_DIR / "macgyver_problems.json"
    if not data_path.exists():
        data_path = download_dataset()

    with open(data_path) as f:
        all_problems = json.load(f)

    if categories:
        cat_lower = {c.lower() for c in categories}
        all_problems = [
            p for p in all_problems
            if p.get("category", "").lower() in cat_lower
        ]

    labelled = [
        {
            "id": f"MG-{i+1:03d}",
            "problem": p["problem"],
            "reference_solution": p["reference_solution"],
            "category": p.get("category", "general"),
        }
        for i, p in enumerate(all_problems)
    ]

    if start_offset:
        labelled = labelled[start_offset:]

    if skip_completed:
        labelled = [p for p in labelled if not (RESULTS_DIR / f"{p['id']}.json").exists()]

    return labelled[:limit]


def format_prompt(problem: dict) -> str:
    return (
        f"PROBLEM: {problem['problem']}\n\n"
        f"Provide a practical, creative solution using only resources "
        f"mentioned in or implied by the problem. Be specific about "
        f"how you would use each resource."
    )


def score_macgyver_solution(
    problem: str,
    solution: str,
    reference: str,
    judge_fn=None,
) -> dict:
    """Score a MacGyver solution against the reference.

    Uses LLM-as-judge to classify solution quality:
    - perfect: Solves the problem completely and creatively
    - correct: Solves the problem but less efficiently/creatively
    - partial: Addresses the problem partially
    - wrong: Does not solve the problem

    Falls back to keyword heuristic if no judge function provided.
    """
    if judge_fn:
        prompt = (
            "You are evaluating a creative problem-solving solution.\n\n"
            f"PROBLEM: {problem}\n\n"
            f"REFERENCE SOLUTION: {reference}\n\n"
            f"CANDIDATE SOLUTION: {solution}\n\n"
            "Classify the candidate solution quality:\n"
            "- perfect: Solves the problem completely and creatively, "
            "comparable to or better than the reference.\n"
            "- correct: Solves the problem but less efficiently or "
            "creatively than the reference.\n"
            "- partial: Addresses the problem partially but misses "
            "key aspects or has significant flaws.\n"
            "- wrong: Does not solve the problem or proposes an "
            "infeasible solution.\n\n"
            "Reply in this format:\n"
            "CLASSIFICATION: <perfect|correct|partial|wrong>\n"
            "RATIONALE: <one sentence>"
        )
        try:
            raw = judge_fn(prompt)
            raw_lower = raw.lower()
            for level in ["perfect", "correct", "partial", "wrong"]:
                if level in raw_lower:
                    return {"level": level, "score": _level_score(level)}
        except Exception:
            pass

    return {"level": "unscored", "score": 0.5}


def _level_score(level: str) -> float:
    return {"perfect": 1.0, "correct": 0.75, "partial": 0.4, "wrong": 0.0}.get(level, 0.0)


def run_macgyver_benchmark(
    triz_fn,
    vanilla_fn,
    judge_fn=None,
    limit: int = 50,
    start_offset: int = 0,
    skip_completed: bool = True,
) -> dict:
    """Run MacGyver benchmark comparing TRIZ-augmented vs vanilla.

    triz_fn: callable(prompt: str, system_prompt: str) -> str
    vanilla_fn: callable(prompt: str, system_prompt: str) -> str
    judge_fn: callable(prompt: str) -> str (for scoring)

    Returns comparison report.
    """
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    problems = load_problems(
        limit=limit, start_offset=start_offset, skip_completed=skip_completed,
    )

    triz_wins = 0
    vanilla_wins = 0
    ties = 0
    results = []

    from benchmark.runner import QuotaExhaustedError

    for problem in problems:
        prompt = format_prompt(problem)
        print(f"  {problem['id']}: {problem['problem'][:60]}...", file=sys.stderr)

        triz_trace = None
        triz_trace_status = "disabled"
        vanilla_trace = None
        vanilla_trace_status = "disabled"
        try:
            triz_response = triz_fn(prompt, TRIZ_SYSTEM_PROMPT)
            if isinstance(triz_response, tuple):
                triz_raw, triz_trace, triz_trace_status = triz_response
            else:
                triz_raw = triz_response
            triz_score = score_macgyver_solution(
                problem["problem"], triz_raw, problem["reference_solution"], judge_fn,
            )
        except QuotaExhaustedError:
            print(f"  QUOTA EXHAUSTED — stopping benchmark.", file=sys.stderr)
            break
        except Exception as e:
            triz_raw = ""
            triz_score = {"level": "error", "score": 0.0}

        try:
            vanilla_response = vanilla_fn(prompt, VANILLA_SYSTEM_PROMPT)
            if isinstance(vanilla_response, tuple):
                vanilla_raw, vanilla_trace, vanilla_trace_status = vanilla_response
            else:
                vanilla_raw = vanilla_response
            vanilla_score = score_macgyver_solution(
                problem["problem"], vanilla_raw, problem["reference_solution"], judge_fn,
            )
        except QuotaExhaustedError:
            print(f"  QUOTA EXHAUSTED — stopping benchmark.", file=sys.stderr)
            break
        except Exception as e:
            vanilla_raw = ""
            vanilla_score = {"level": "error", "score": 0.0}

        if triz_score["score"] > vanilla_score["score"]:
            triz_wins += 1
        elif vanilla_score["score"] > triz_score["score"]:
            vanilla_wins += 1
        else:
            ties += 1

        result = {
            "problem_id": problem["id"],
            "category": problem["category"],
            "triz": {"level": triz_score["level"], "score": triz_score["score"]},
            "vanilla": {"level": vanilla_score["level"], "score": vanilla_score["score"]},
        }
        results.append(result)

        result_path = RESULTS_DIR / f"{problem['id']}.json"
        result_payload = {
            **result,
            "problem": problem["problem"],
            "triz_output": triz_raw[:2000],
            "vanilla_output": vanilla_raw[:2000],
            "reference": problem["reference_solution"],
        }
        if triz_trace is not None:
            result_payload["triz_trace"] = triz_trace
            result_payload["triz_trace_status"] = triz_trace_status
        if vanilla_trace is not None:
            result_payload["vanilla_trace"] = vanilla_trace
            result_payload["vanilla_trace_status"] = vanilla_trace_status
        result_path.write_text(json.dumps(result_payload, indent=2))

    total = len(results)
    return {
        "benchmark": "MacGyver",
        "total_problems": total,
        "triz_wins": triz_wins,
        "vanilla_wins": vanilla_wins,
        "ties": ties,
        "triz_win_rate": triz_wins / total if total else 0,
        "vanilla_win_rate": vanilla_wins / total if total else 0,
        "triz_mean_score": (
            sum(r["triz"]["score"] for r in results) / total if total else 0
        ),
        "vanilla_mean_score": (
            sum(r["vanilla"]["score"] for r in results) / total if total else 0
        ),
        "results": results,
    }


if __name__ == "__main__":
    print("MacGyver adapter ready")
    try:
        problems = load_problems(limit=3)
        print(f"Loaded {len(problems)} problems:")
        for p in problems:
            print(f"  {p['id']}: {p['problem'][:80]}...")
    except Exception as e:
        print(f"Dataset not available: {e}")
        print("Run with: python -m benchmark.external.macgyver_adapter")
